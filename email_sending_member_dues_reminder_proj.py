#! python 2.7
# sendDuesReminder.py - Sends emails based on payment status in spreadsheet

import openpyxl, smtplib, sys


#TODO: Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook('email_excelFile.xlsx')     # Then we get Sheet 1 and store the resulting Worksheet object in sheet
sheet = wb.get_sheet_by_name('Sheet1')                  # Now that we have a Worksheet object, we can access rows, columns, and cells. We store the highest column in lastCol

lastCol = sheet.max_column                              # we then use row number 1 and lastCol to access the cell that should hold the most recent month.
latestMonth = sheet.cell(row=1, column=lastCol).value   # We get the value in this cell and store it in latestMonth


#TODO: Check each member's payment status.
unpaidMembers = {}
for r in range(2, sheet.max_row + 1):         # This code sets up an empty dictionary unpaidMembers and then loops through all the rows after the first
    payment = sheet.cell(row=r, column=lastCol).value   # For each row, the value in the most recent column is stored in payment
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value        # If payment is not equal to 'paid', then the value of the first column is stored in name
        email = sheet.cell(row=r, column=2).value       # the value of the second column is stored in email
        unpaidMembers[name] = email


#TODO: Log in to email account.
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login(' ksoh512@gmail.com ', sys.argv[1])



#TODO: Send out reminder emails.
for name, email in unpaidMembers.items():                                                       # This code loops through the names and emails in unpaidMembers. For each member who hasn’t paid, we customize a message with the latest month and the member’s name, and store the message in body
    body = "Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not
        paid dues for %s. Please make this payment as soon as possible. Thank you!'" %
        (latestMonth, name, latestMonth)
    print('Sending email to %s...' % email)                                                     # We print output saying that we’re sending an email to this member’s email address
    sendmailStatus = smtpObj.sendmail('my_email_address@gmail.com', email, body)                # Then we call sendmail(), passing it the from address and the customized message

    if sendmailStatus != {}:                                                                    # Remember that the sendmail() method will return a nonempty dictionary value if the SMTP server reported an error sending that particular email. The last part of the for loop at ❹ checks if the returned dictionary is nonempty, and if it is, prints the recipient’s email address and the returned dictionary.
        print('There was a problem sending email to %s: %s' % (email,
        sendmailStatus))
smtpObj.quit()
